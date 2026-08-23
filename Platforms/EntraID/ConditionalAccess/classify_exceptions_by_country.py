"""
classify_exceptions_by_country.py
Splits/tags the "All Exceptions MFA" sheet of an MFA exception report by country.

Why not just usageLocation:
  Many rows are service/shared accounts with no reliable usageLocation, so we cross-check
  two independent signals instead:
    1) UPN domain suffix  -> "home" country, where the domain is country-specific
    2) Countries column   -> actual sign-in geography already collected in this report
  Central multi-country domains (the main corporate domain, the .onmicrosoft.com
  domain, and subdomains of either) cannot be resolved by domain alone, so sign-in
  country is used instead. If you have a usageLocation export (UPN + usageLocation
  columns), pass it with --usage-location to add it as a third cross-check column.

Domain data:
  The domain -> country mapping is NOT hardcoded here. It is read from a shared
  PowerShell data file (see --domain-map-path), which ships with placeholder
  contoso.* domains - replace them with your own before running.

Usage:
  python classify_exceptions_by_country.py <input.xlsx> [--sheet "All Exceptions MFA"]
      [--usage-location lookup.csv] [--domain-map-path ../../_Shared/Data/domain-country-map.psd1]
      [--out output.xlsx]
"""
import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter

# ---------------- CONFIG ----------------
# Shared domain/country data, also consumed by the PowerShell scripts in this repo.
DEFAULT_DOMAIN_MAP_PATH = (
    Path(__file__).resolve().parents[2] / "_Shared" / "Data" / "domain-country-map.psd1"
)

# Populated from the data file by load_domain_map().
DOMAIN_COUNTRY_MAP: dict[str, str] = {}
CENTRAL_DOMAINS: set[str] = set()
# -----------------------------------------


def _strip_psd1_comments(text: str) -> str:
    """Drop '#' comments that are not inside a single-quoted string."""
    out = []
    for line in text.splitlines():
        in_quote = False
        cut = len(line)
        for i, ch in enumerate(line):
            if ch == "'":
                in_quote = not in_quote
            elif ch == "#" and not in_quote:
                cut = i
                break
        out.append(line[:cut])
    return "\n".join(out)


def _psd1_block(text: str, name: str, opener: str) -> str:
    """Return the body of a top-level `name = @{...}` / `name = @(...)` block."""
    closer = "}" if opener == "{" else ")"
    m = re.search(r"\b" + re.escape(name) + r"\s*=\s*@" + re.escape(opener), text)
    if not m:
        raise ValueError(f"Section '{name}' not found in the domain map file")
    depth, start = 1, m.end()
    for i in range(start, len(text)):
        if text[i] == opener:
            depth += 1
        elif text[i] == closer:
            depth -= 1
            if depth == 0:
                return text[start:i]
    raise ValueError(f"Section '{name}' is not closed in the domain map file")


def load_domain_map(path: Path) -> None:
    """Read DomainToCountry / CentralDomains out of the shared .psd1 data file.

    Deliberately a small reader for the subset of PSD1 that file uses (flat
    hashtables and string arrays, single-quoted) - it does not evaluate PowerShell.
    """
    global DOMAIN_COUNTRY_MAP, CENTRAL_DOMAINS
    text = _strip_psd1_comments(Path(path).read_text(encoding="utf-8-sig"))

    pairs = re.findall(
        r"'([^']+)'\s*=\s*'([^']*)'", _psd1_block(text, "DomainToCountry", "{")
    )
    DOMAIN_COUNTRY_MAP = {k.lower(): v for k, v in pairs}

    CENTRAL_DOMAINS = {
        d.lower() for d in re.findall(r"'([^']+)'", _psd1_block(text, "CentralDomains", "("))
    }

    if not DOMAIN_COUNTRY_MAP or not CENTRAL_DOMAINS:
        raise ValueError(f"Domain map {path} produced no usable entries")


def domain_country(domain: str) -> tuple[str, str]:
    """Return (bucket, country_or_None). bucket in {mapped, central, unknown}."""
    if not domain or "@" in domain:
        return "unknown", None
    d = domain.lower()
    if d in DOMAIN_COUNTRY_MAP:
        return "mapped", DOMAIN_COUNTRY_MAP[d]
    if d in CENTRAL_DOMAINS or any(d.endswith("." + c) for c in CENTRAL_DOMAINS):
        return "central", None
    return "unknown", None


def parse_signin_countries(val) -> list[str]:
    if pd.isna(val) or not str(val).strip():
        return []
    return [c.strip() for c in re.split(r"[;,]", str(val)) if c.strip()]


def classify(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Domain"] = df["UserPrincipalName"].astype(str).str.split("@").str[-1]
    df["SignIn Countries List"] = df["Countries"].apply(parse_signin_countries)

    domain_bucket, domain_ctry, assigned, method, flag = [], [], [], [], []

    for _, row in df.iterrows():
        bucket, ctry = domain_country(row["Domain"])
        signin = row["SignIn Countries List"]
        domain_bucket.append(bucket)
        domain_ctry.append(ctry or "")

        if bucket == "mapped":
            a, m = ctry, "Domain (unambiguous)"
            if not signin:
                f = "No sign-in data - domain only"
            elif ctry in signin and len(signin) == 1:
                f = "Match"
            elif ctry in signin and len(signin) > 1:
                f = "Match (also seen from other countries)"
            else:
                f = f"Mismatch - sign-in shows {','.join(signin)}"
        elif bucket == "central":
            if len(signin) == 1:
                a, m, f = signin[0], "Sign-in (central domain, single country)", "OK"
            elif len(signin) > 1:
                a, m, f = "MULTIPLE: " + ",".join(signin), "Sign-in (central domain, multiple countries)", "Manual review - multiple sign-in countries"
            else:
                a, m, f = "UNKNOWN", "No signal available", "Needs usageLocation lookup"
        else:
            if len(signin) == 1:
                a, m, f = signin[0], "Sign-in (unresolvable domain/object)", "OK - domain unresolved, sign-in used"
            elif len(signin) > 1:
                a, m, f = "MULTIPLE: " + ",".join(signin), "Sign-in (unresolvable domain/object)", "Manual review - multiple sign-in countries"
            else:
                a, m, f = "UNKNOWN", "No signal available", "Needs usageLocation lookup"

        assigned.append(a)
        method.append(m)
        flag.append(f)

    df["Domain Bucket"] = domain_bucket
    df["Domain Country"] = domain_ctry
    df["Assigned Country"] = assigned
    df["Assignment Method"] = method
    df["Country Flag"] = flag
    df.drop(columns=["SignIn Countries List"], inplace=True)
    return df


def merge_usage_location(df: pd.DataFrame, path: str) -> pd.DataFrame:
    ul = pd.read_csv(path)
    ul.columns = [c.strip() for c in ul.columns]
    key = "UserPrincipalName" if "UserPrincipalName" in ul.columns else ul.columns[0]
    loc_col = "usageLocation" if "usageLocation" in ul.columns else ul.columns[1]
    ul = ul[[key, loc_col]].rename(columns={key: "UserPrincipalName", loc_col: "usageLocation"})
    df = df.merge(ul, on="UserPrincipalName", how="left")

    def resolve(r):
        has_loc = not pd.isna(r.get("usageLocation")) and str(r.get("usageLocation")).strip() != ""
        assigned = r["Assigned Country"] or ""

        if assigned.startswith("UNKNOWN"):
            if has_loc:
                return pd.Series([r["usageLocation"], "usageLocation (Graph) - no sign-in data", "Resolved via usageLocation"])
            return pd.Series([assigned, r["Assignment Method"], "Needs manual investigation - no usageLocation, no sign-in data"])

        if assigned.startswith("MULTIPLE"):
            signin_list = assigned.replace("MULTIPLE: ", "").split(",")
            if has_loc and r["usageLocation"] in signin_list:
                return pd.Series([r["usageLocation"], "usageLocation (Graph) - matched one of several sign-in countries", "Resolved via usageLocation"])
            elif has_loc:
                return pd.Series([assigned, r["Assignment Method"], f"Manual review - usageLocation={r['usageLocation']} not in sign-in list ({assigned})"])
            return pd.Series([assigned, r["Assignment Method"], r["Country Flag"]])

        # already resolved via domain or single sign-in country - just check agreement
        if not has_loc:
            return pd.Series([assigned, r["Assignment Method"], r["Country Flag"]])
        if r["usageLocation"] == assigned:
            return pd.Series([assigned, r["Assignment Method"], "Match (confirmed by usageLocation)"])
        return pd.Series([assigned, r["Assignment Method"], f"Mismatch - usageLocation={r['usageLocation']} vs assigned={assigned}"])

    df[["Assigned Country", "Assignment Method", "Country Flag"]] = df.apply(resolve, axis=1)
    return df


def write_output(df: pd.DataFrame, out_path: str):
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        df.to_excel(xw, sheet_name="All Exceptions MFA (by country)", index=False)
        summary = (
            df.groupby(["Assigned Country", "Tier"], dropna=False).size().unstack(fill_value=0)
        )
        summary["Total"] = summary.sum(axis=1)
        summary = summary.sort_values("Total", ascending=False)
        summary.to_excel(xw, sheet_name="Country Summary")
        review = df[df["Country Flag"].str.contains("Manual review|Mismatch|Needs usageLocation", na=False)]
        review.to_excel(xw, sheet_name="Needs Review", index=False)

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    for ws in wb.worksheets:
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 60)
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("--sheet", default="All Exceptions MFA")
    ap.add_argument("--usage-location", default=None, help="Optional CSV with UserPrincipalName,usageLocation columns")
    ap.add_argument(
        "--domain-map-path",
        default=str(DEFAULT_DOMAIN_MAP_PATH),
        help="PSD1 data file holding DomainToCountry / CentralDomains "
             "(default: the sample map in Platforms/_Shared/Data)",
    )
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    out = args.out or args.input.replace(".xlsx", "_by_country.xlsx")

    load_domain_map(Path(args.domain_map_path))
    print(
        f"Domain map: {args.domain_map_path} "
        f"({len(DOMAIN_COUNTRY_MAP)} domains, {len(CENTRAL_DOMAINS)} central)"
    )

    df = pd.read_excel(args.input, sheet_name=args.sheet)
    df = classify(df)
    if args.usage_location:
        df = merge_usage_location(df, args.usage_location)

    write_output(df, out)

    total = len(df)
    review_n = df["Country Flag"].str.contains("Manual review|Mismatch|Needs usageLocation", na=False).sum()
    print(f"Classified {total} rows. Flagged for review: {review_n}. Output: {out}")


if __name__ == "__main__":
    main()